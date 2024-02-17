from bs4 import BeautifulSoup
import requests as req
import openpyxl


def parse_7745(category, excel_table):
    url = f"https://7745.by/catalog/{category}?limit=1000"
    page = req.get(url)

    allLaptops = []
    filteredLaptops = []

    soup = BeautifulSoup(page.text, "html.parser")  # превращаю страницу сайта в объект, с которым можно работать
    allLaptops = soup.findAll('div', class_='catalog-item__wrapper')  # находим все div с классом catalog-item__wrapper (карточки товаров)

    wb = openpyxl.load_workbook(excel_table)  # открываю существующую таблицу excel
    wb.create_sheet(category.capitalize())  # создаю в ней новый листик
    sheet = wb[category.capitalize()]  # нахожу этот листик в таблице

    sheet.append(['Название товара', 'Стоимость', 'Артикул'])  # добавление заголовков таблицы
    for laptop in allLaptops:
        name = laptop.find('a', class_='item-block_name').text
        price = laptop.find('div', class_='item-block_main-price-block').find('span').text.strip()[:9]
        art = laptop.find('div', class_='product__actions-block--right').find('span').text.strip()
        row = [name, price, art]
        sheet.append(row)

    wb.save(excel_table)

wb = openpyxl.Workbook()
wb.save('parse_7745.xlsx')

cat = ['noutbuki', 'smartfony', 'televizory', 'ccp-umnyy-dom-alisa-yandeks']
for categ in cat:
    parse_7745(categ, 'parse_7745.xlsx')

